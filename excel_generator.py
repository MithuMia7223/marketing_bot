import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_leads_excel(leads: list, query: str) -> str:
    """Generates a beautifully styled Excel file containing the leads.
    Returns the absolute path to the generated file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "B2B Leads"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Headers definition
    headers = ["ID", "Company Name", "Website", "Email", "Phone", "WhatsApp Link"]
    ws.append(headers)
    
    # Style definitions
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # Borders
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Style Header Row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align if col_num in [1, 5, 6] else left_align
        cell.border = thin_border
        
    # Append Data
    data_font = Font(name=font_family, size=10)
    for lead in leads:
        row_data = [
            lead["id"],
            lead["name"],
            lead["website"],
            lead["email"],
            lead["phone"],
            lead["whatsapp"]
        ]
        ws.append(row_data)
        
        # Style Data Row
        current_row = ws.max_row
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.font = data_font
            cell.alignment = center_align if col_num in [1, 5] else left_align
            cell.border = thin_border
            
            # Format links
            if col_num in [3, 6] and cell.value and cell.value != "Not found":
                cell.hyperlink = cell.value
                cell.font = Font(name=font_family, size=10, color="0000FF", underline="single")

    # Auto-adjust column widths based on content length
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        # Apply width with a small padding
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Enable Autofilters
    ws.auto_filter.ref = f"A1:F{ws.max_row}"
    
    # Ensure exports folder exists
    export_dir = "exports"
    if not os.path.exists(export_dir):
        os.makedirs(export_dir)
        
    # Clean filename
    clean_query = "".join([c if c.isalnum() else "_" for c in query]).strip("_")
    filename = f"exports/Leads_{clean_query}.xlsx"
    
    wb.save(filename)
    return filename
